"""
Exportação: PDF, CSV, Excel.
"""
import io
from html import escape
from typing import Any, List, Optional

from fastapi.responses import StreamingResponse


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _format_br_number(v: Any) -> str:
    """Formata número para exibição (pt-BR: 1.234,56)."""
    if v is None:
        return "0"
    try:
        n = float(v)
        s = f"{n:,.2f}"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return str(v)


def export_csv(rows: List[dict], filename: str = "export.csv") -> StreamingResponse:
    """Gera CSV a partir de lista de dicts."""
    if not rows:
        output = "No data\n"
    else:
        keys = list(rows[0].keys())
        lines = [",".join(f'"{_safe_str(k)}"' for k in keys)]
        for r in rows:
            lines.append(",".join(f'"{_safe_str(r.get(k))}"' for k in keys))
        output = "\n".join(lines)
    return StreamingResponse(
        iter([output.encode("utf-8-sig")]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def export_excel(rows: List[dict], filename: str = "export.xlsx") -> StreamingResponse:
    """Gera Excel com openpyxl."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        if rows:
            keys = list(rows[0].keys())
            for col, k in enumerate(keys, 1):
                ws.cell(row=1, column=col, value=k)
            for row_idx, r in enumerate(rows, 2):
                for col_idx, k in enumerate(keys, 1):
                    ws.cell(row=row_idx, column=col_idx, value=r.get(k))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"No data or openpyxl error"]),
            media_type="text/plain",
            status_code=500,
        )


def export_dashboard_excel(
    title: str,
    rows: List[dict],
    filename: str = "dashboard-stock-expiry.xlsx",
    subtitle: Optional[str] = None,
) -> StreamingResponse:
    """Gera Excel do dashboard com layout: título, filtros (opcional), tabela com cabeçalho verde e colunas ajustadas."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Estoque a vencer"

        header_fill = PatternFill(start_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        align_right = Alignment(horizontal="right")
        wrap_align = Alignment(wrap_text=True, vertical="center")

        row_num = 1
        ws.cell(row=row_num, column=1, value=title)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 1
        if subtitle and subtitle.strip():
            ws.cell(row=row_num, column=1, value=subtitle)
            ws.cell(row=row_num, column=1).font = Font(size=9, color="555555")
            row_num += 1
        row_num += 1

        headers = ["Cód. material", "Material", "Almoxarifado", "Qtd", "Valor total (R$)", "Validade"]
        keys = ["material_code", "material_name", "warehouse", "quantity", "total_value", "expiry_date"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_num += 1

        for r in rows[:2000]:
            for col_idx, key in enumerate(keys, 1):
                val = r.get(key)
                if val is None:
                    val = ""
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.border = border
                if col_idx in (4, 5):
                    c.alignment = align_right
                else:
                    c.alignment = wrap_align
            row_num += 1

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 38
        ws.column_dimensions["D"].width = 8
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"Excel generation error"]),
            media_type="text/plain",
            status_code=500,
        )


def export_pdf_simple(
    title: str,
    rows: List[dict],
    filename: str = "report.pdf",
    subtitle: Optional[str] = None,
) -> StreamingResponse:
    """Gera PDF em orientação horizontal (paisagem) com tabela legível. Opcional: subtitle (ex.: filtros aplicados)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.2 * cm,
            rightMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )
        story = []
        styles = getSampleStyleSheet()
        title_style = styles["Title"]
        story.append(Paragraph(title, title_style))
        if subtitle and subtitle.strip():
            sub_style = styles["Normal"].__class__(name="Subtitle", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#555"))
            story.append(Paragraph(escape(subtitle.strip()), sub_style))

        if not rows:
            data = [["Nenhum dado para exibir."]]
            col_widths = [20 * cm]
        else:
            # Estilos para células: texto com quebra (Paragraph) e alinhamento
            normal = getSampleStyleSheet()["Normal"]
            table_text = normal.__class__(name="TableText", parent=normal, fontSize=8, leading=9)
            table_header = normal.__class__(name="TableHeader", parent=normal, fontSize=9, alignment=TA_CENTER)
            table_right = normal.__class__(name="TableRight", parent=normal, fontSize=8, leading=9, alignment=TA_RIGHT)

            headers = ["Cód. material", "Material", "Almoxarifado", "Qtd", "Valor total (R$)", "Validade"]
            keys = ["material_code", "material_name", "warehouse", "quantity", "total_value", "expiry_date"]
            # Cabeçalho como Paragraphs para consistência
            data = [[Paragraph(escape(h), table_header) for h in headers]]
            for r in rows[:500]:
                cod = _safe_str(r.get("material_code"))
                material = _safe_str(r.get("material_name"))
                almox = _safe_str(r.get("warehouse"))
                qtd = _safe_str(r.get("quantity"))
                valor = _safe_str(r.get("total_value"))
                validade = _safe_str(r.get("expiry_date"))
                data.append([
                    Paragraph(escape(cod) or "—", table_text),
                    Paragraph(escape(material) or "—", table_text),
                    Paragraph(escape(almox) or "—", table_text),
                    Paragraph(escape(qtd), table_right),
                    Paragraph(escape(valor), table_right),
                    Paragraph(escape(validade), table_right),
                ])
            # Larguras em paisagem: mais espaço para Material e Almoxarifado
            col_widths = [2.2 * cm, 6.5 * cm, 5.5 * cm, 1.5 * cm, 2.8 * cm, 2.5 * cm]

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8BC547")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.yellow),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#faf8f5")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (2, -1), "LEFT"),
            ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"PDF generation error"]),
            media_type="text/plain",
            status_code=500,
        )


def _format_validity_dd_mm_yyyy(d: Any) -> str:
    """Formata data para dd-mm-yyyy (regra tela itens vencidos)."""
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    try:
        from datetime import date
        if isinstance(d, date):
            return d.strftime("%d-%m-%Y")
    except Exception:
        pass
    return str(d)


def _format_validity_mm_yyyy(d: Any) -> str:
    """Formata data para MM/YYYY (coluna Mês - Detalhes dos Itens Vencidos)."""
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    try:
        from datetime import date
        if isinstance(d, date):
            return d.strftime("%m/%Y")
    except Exception:
        pass
    return str(d)


def _format_validity_dd_mm_yyyy_slash(d: Any) -> str:
    """Formata data para dd/mm/yyyy (ex.: 31/07/2026)."""
    if d is None:
        return ""
    if isinstance(d, str):
        s = d.strip()
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            # ISO YYYY-MM-DD
            return f"{s[8:10]}/{s[5:7]}/{s[0:4]}"
        return s
    try:
        from datetime import date
        if isinstance(d, date):
            return d.strftime("%d/%m/%Y")
    except Exception:
        pass
    return str(d)


def export_predictive_pdf(
    title: str,
    rows: List[dict],
    filename: str = "analise-preditiva.pdf",
    subtitle: Optional[str] = None,
) -> StreamingResponse:
    """Gera PDF da análise preditiva: orientação horizontal, dados formatados (moeda pt-BR, datas dd/mm/yyyy, risco)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.0 * cm,
            rightMargin=1.0 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )
        story = []
        styles = getSampleStyleSheet()
        story.append(Paragraph(escape(title), styles["Title"]))
        if subtitle and subtitle.strip():
            sub_style = styles["Normal"].__class__(
                name="Subtitle", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#555")
            )
            story.append(Paragraph(escape(subtitle.strip()), sub_style))

        if not rows:
            data = [["Nenhum dado para exibir."]]
            col_widths = [20 * cm]
        else:
            normal = getSampleStyleSheet()["Normal"]
            table_text = normal.__class__(name="TableText", parent=normal, fontSize=7, leading=8)
            table_header = normal.__class__(name="TableHeader", parent=normal, fontSize=8, alignment=TA_CENTER)
            table_right = normal.__class__(name="TableRight", parent=normal, fontSize=7, leading=8, alignment=TA_RIGHT)
            headers = [
                "Material", "Grupo", "Almoxarifado", "Lote", "Validade", "Dias p/ vencer",
                "Qtd. disp.", "Valor unit.", "Valor Total", "Cons. méd/mês",
                "Mes/ano último consumo", "Qtde último consumo", "Risco de perda",
                "Previsão Perda", "Valor est. perda (R$)",
            ]
            data = [[Paragraph(escape(h), table_header) for h in headers]]
            for r in rows[:500]:
                material = _safe_str(r.get("material_name") or r.get("material_code") or "—")
                if len(material) > 120:
                    material = material[:117] + "..."
                data.append([
                    Paragraph(escape(material) or "—", table_text),
                    Paragraph(escape(_safe_str(r.get("material_group")) or "—"), table_text),
                    Paragraph(escape(_safe_str(r.get("almoxarifado")) or "—"), table_text),
                    Paragraph(escape(_safe_str(r.get("lote")) or "—"), table_text),
                    Paragraph(_format_validity_dd_mm_yyyy_slash(r.get("validity")), table_right),
                    Paragraph(_safe_str(r.get("days_until_expiry")), table_right),
                    Paragraph(_format_br_number(r.get("quantity")), table_right),
                    Paragraph(_format_br_number(r.get("unit_value")), table_right),
                    Paragraph(_format_br_number(r.get("total_value")), table_right),
                    Paragraph(_format_br_number(r.get("avg_monthly_consumption")), table_right),
                    Paragraph(_safe_str(r.get("last_consumption_mesano")) or "—", table_right),
                    Paragraph(_format_br_number(r.get("qtde_ultimo_consumo")), table_right),
                    Paragraph(escape(_safe_str(r.get("risk")) or "—"), table_text),
                    Paragraph(_format_br_number(r.get("predicted_loss_quantity")), table_right),
                    Paragraph(_format_br_number(r.get("estimated_loss")), table_right),
                ])
            col_widths = [
                4.0 * cm, 2.2 * cm, 2.8 * cm, 1.4 * cm, 1.6 * cm, 1.4 * cm,
                1.2 * cm, 1.6 * cm, 1.8 * cm, 1.4 * cm,
                1.6 * cm, 1.2 * cm,  # Mes/ano último consumo, Qtde último consumo
                2.0 * cm, 1.2 * cm, 2.0 * cm,
            ]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#faf8f5")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (3, -1), "LEFT"),
            ("ALIGN", (4, 0), (-1, -1), "RIGHT"),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"PDF generation error"]),
            media_type="text/plain",
            status_code=500,
        )


def export_expired_pdf(
    title: str,
    rows: List[dict],
    filename: str = "itens-vencidos.pdf",
    subtitle: Optional[str] = None,
) -> StreamingResponse:
    """PDF formatado: Detalhes dos Itens Vencidos (Material, Mês, Qtd, Valor unit., Valor total, Grupo, Almoxarifado, Status).
    Tabela em A4 paisagem com coluna Material alargada para melhor leitura das descrições longas."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=landscape(A4),
            leftMargin=1.0 * cm,
            rightMargin=1.0 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm,
        )
        story = []
        styles = getSampleStyleSheet()
        story.append(Paragraph(title, styles["Title"]))
        if subtitle and subtitle.strip():
            sub_style = styles["Normal"].__class__(name="Subtitle", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#555"))
            story.append(Paragraph(escape(subtitle.strip()), sub_style))

        if not rows:
            data = [["Nenhum dado para exibir."]]
            col_widths = [20 * cm]
        else:
            normal = getSampleStyleSheet()["Normal"]
            table_text = normal.__class__(name="TableText", parent=normal, fontSize=7, leading=8)
            table_header = normal.__class__(name="TableHeader", parent=normal, fontSize=8, alignment=TA_CENTER)
            table_right = normal.__class__(name="TableRight", parent=normal, fontSize=7, leading=8, alignment=TA_RIGHT)
            headers = ["Material", "Mês", "Qtd", "Valor unit.", "Valor total", "Grupo", "Almoxarifado", "Status"]
            data = [[Paragraph(escape(h), table_header) for h in headers]]
            for r in rows[:500]:
                data.append([
                    Paragraph(escape(_safe_str(r.get("material_name")) or "—"), table_text),
                    Paragraph(escape(_safe_str(r.get("validity"))), table_right),
                    Paragraph(_format_br_number(r.get("quantity")), table_right),
                    Paragraph(_format_br_number(r.get("unit_value")), table_right),
                    Paragraph(_format_br_number(r.get("total_value")), table_right),
                    Paragraph(escape(_safe_str(r.get("group")) or "—"), table_text),
                    Paragraph(escape(_safe_str(r.get("warehouse")) or "—"), table_text),
                    Paragraph(escape(_safe_str(r.get("status")) or "VENCIDO"), table_text),
                ])
            # Larguras pensadas para A4 paisagem (~27,7 cm úteis): Material bem largo para menos quebras de linha
            col_widths = [
                8.5 * cm,   # Material (descrições longas; antes 2.5)
                2.0 * cm,   # Mês
                1.2 * cm,   # Qtd
                2.0 * cm,   # Valor unit.
                2.2 * cm,   # Valor total
                4.0 * cm,   # Grupo
                5.0 * cm,   # Almoxarifado
                1.5 * cm,   # Status
            ]
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 1), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#faf8f5")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (4, -1), "RIGHT"),
            ("ALIGN", (5, 0), (6, -1), "LEFT"),
            ("ALIGN", (7, 0), (7, -1), "CENTER"),
        ]))
        story.append(t)
        doc.build(story)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"PDF generation error"]),
            media_type="text/plain",
            status_code=500,
        )


def export_expired_excel(
    title: str,
    rows: List[dict],
    filename: str = "itens-vencidos.xlsx",
    subtitle: Optional[str] = None,
) -> StreamingResponse:
    """Excel formatado: Detalhes dos Itens Vencidos (colunas completas)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Itens vencidos"

        header_fill = PatternFill(start_color="2E7D32", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_side = Side(style="thin", color="000000")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        align_right = Alignment(horizontal="right")
        wrap_align = Alignment(wrap_text=True, vertical="center")

        row_num = 1
        ws.cell(row=row_num, column=1, value=title)
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
        row_num += 1
        if subtitle and subtitle.strip():
            ws.cell(row=row_num, column=1, value=subtitle)
            ws.cell(row=row_num, column=1).font = Font(size=9, color="555555")
            row_num += 1
        row_num += 1

        headers = ["Material", "Mês", "Qtd", "Valor unit.", "Valor total", "Grupo", "Almoxarifado", "Status"]
        keys = ["material_name", "validity", "quantity", "unit_value", "total_value", "group", "warehouse", "status"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=row_num, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        row_num += 1

        for r in rows[:2000]:
            for col_idx, key in enumerate(keys, 1):
                val = r.get(key)
                if val is None:
                    val = ""
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.border = border
                if col_idx in (3, 4, 5):  # Qtd, Valor unit., Valor total
                    c.alignment = align_right
                else:
                    c.alignment = wrap_align
            row_num += 1

        for col_letter, width in [("A", 22), ("B", 12), ("C", 8), ("D", 12), ("E", 14), ("F", 28), ("G", 38), ("H", 10)]:
            ws.column_dimensions[col_letter].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception:
        return StreamingResponse(
            iter([b"Excel generation error"]),
            media_type="text/plain",
            status_code=500,
        )
